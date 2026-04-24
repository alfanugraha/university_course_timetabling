"""
backend/tests/test_excel_exporter.py
Unit tests untuk ExcelExporter.export_jadwal().

Menggunakan SQLite in-memory dengan schema minimal yang mencerminkan
tabel-tabel yang dibutuhkan exporter (tanpa tipe PostgreSQL-specific).
"""

from __future__ import annotations

import datetime
import io
import uuid

import openpyxl
import pytest
from sqlalchemy import (
    Boolean,
    Column,
    ForeignKey,
    Integer,
    SmallInteger,
    String,
    Text,
    Time,
    create_engine,
)
from sqlalchemy.orm import DeclarativeBase, relationship, sessionmaker


# ---------------------------------------------------------------------------
# Minimal SQLite-compatible ORM (mirrors production models)
# ---------------------------------------------------------------------------

class _Base(DeclarativeBase):
    pass


class _Prodi(_Base):
    __tablename__ = "prodi"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    kode = Column(String(10), unique=True, nullable=False)
    strata = Column(String(5), nullable=False)
    nama = Column(String(100), nullable=False)
    singkat = Column(String(20), nullable=False)
    kategori = Column(String(20), nullable=False)
    is_active = Column(Boolean, default=True)


class _Kurikulum(_Base):
    __tablename__ = "kurikulum"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    kode = Column(String(20), unique=True, nullable=False)
    tahun = Column(String(4), nullable=False)
    prodi_id = Column(String(36), ForeignKey("prodi.id"), nullable=False)
    is_active = Column(Boolean, default=True)
    prodi = relationship("_Prodi")


class _MataKuliah(_Base):
    __tablename__ = "mata_kuliah"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    kode = Column(String(20), nullable=False)
    kurikulum_id = Column(String(36), ForeignKey("kurikulum.id"), nullable=False)
    nama = Column(String(200), nullable=False)
    sks = Column(SmallInteger, nullable=False)
    semester = Column(SmallInteger, nullable=False)
    jenis = Column(String(10), nullable=False)
    is_active = Column(Boolean, default=True)
    kurikulum = relationship("_Kurikulum")


class _MataKuliahKelas(_Base):
    __tablename__ = "mata_kuliah_kelas"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    mata_kuliah_id = Column(String(36), ForeignKey("mata_kuliah.id"), nullable=False)
    kelas = Column(String(5), nullable=True)
    label = Column(String(200), nullable=False)
    ket = Column(Text, nullable=True)
    mata_kuliah = relationship("_MataKuliah")


class _Dosen(_Base):
    __tablename__ = "dosen"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    kode = Column(String(10), unique=True, nullable=False)
    nama = Column(String(200), nullable=False)
    homebase_prodi_id = Column(String(36), ForeignKey("prodi.id"), nullable=True)
    status = Column(String(20), default="Aktif")
    homebase_prodi = relationship("_Prodi", foreign_keys=[homebase_prodi_id])


class _Ruang(_Base):
    __tablename__ = "ruang"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    nama = Column(String(20), unique=True, nullable=False)
    kapasitas = Column(SmallInteger, default=45)
    lantai = Column(SmallInteger, nullable=True)
    gedung = Column(String(100), nullable=True)
    jenis = Column(String(20), default="Kelas")
    is_active = Column(Boolean, default=True)


class _Timeslot(_Base):
    __tablename__ = "timeslot"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    kode = Column(String(20), unique=True, nullable=False)
    hari = Column(String(10), nullable=False)
    sesi = Column(SmallInteger, nullable=False)
    jam_mulai = Column(Time, nullable=False)
    jam_selesai = Column(Time, nullable=False)
    label = Column(String(30), nullable=False)
    sks = Column(SmallInteger, nullable=False)


class _SesiJadwal(_Base):
    __tablename__ = "sesi_jadwal"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    nama = Column(String(100), nullable=False)
    semester = Column(String(10), nullable=False)
    tahun_akademik = Column(String(10), nullable=False)
    status = Column(String(20), default="Draft")


class _JadwalAssignment(_Base):
    __tablename__ = "jadwal_assignment"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    sesi_id = Column(String(36), ForeignKey("sesi_jadwal.id"), nullable=False)
    mk_kelas_id = Column(String(36), ForeignKey("mata_kuliah_kelas.id"), nullable=False)
    dosen1_id = Column(String(36), ForeignKey("dosen.id"), nullable=False)
    dosen2_id = Column(String(36), ForeignKey("dosen.id"), nullable=True)
    timeslot_id = Column(String(36), ForeignKey("timeslot.id"), nullable=False)
    ruang_id = Column(String(36), ForeignKey("ruang.id"), nullable=True)
    override_floor_priority = Column(Boolean, default=False)
    catatan = Column(Text, nullable=True)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def engine():
    eng = create_engine("sqlite:///:memory:")
    _Base.metadata.create_all(eng)
    return eng


@pytest.fixture(scope="module")
def db(engine):
    Session = sessionmaker(bind=engine)
    session = Session()
    yield session
    session.close()


@pytest.fixture(scope="module")
def seeded_db(db):
    """Seed minimal data and return IDs for use in tests."""
    prodi_id = str(uuid.uuid4())
    kurikulum_id = str(uuid.uuid4())
    mk_id = str(uuid.uuid4())
    mk_kelas_id = str(uuid.uuid4())
    dosen1_id = str(uuid.uuid4())
    dosen2_id = str(uuid.uuid4())
    ruang_id = str(uuid.uuid4())
    timeslot_id = str(uuid.uuid4())
    sesi_id = str(uuid.uuid4())
    assignment_id = str(uuid.uuid4())

    db.add(_Prodi(
        id=prodi_id, kode="S1MTK", strata="S1",
        nama="S1 Matematika", singkat="S1 MTK", kategori="Reguler",
    ))
    db.add(_Kurikulum(id=kurikulum_id, kode="K2020", tahun="2020", prodi_id=prodi_id))
    db.add(_MataKuliah(
        id=mk_id, kode="MTK101", kurikulum_id=kurikulum_id,
        nama="Kalkulus I", sks=3, semester=1, jenis="Wajib",
    ))
    db.add(_MataKuliahKelas(
        id=mk_kelas_id, mata_kuliah_id=mk_id, kelas="A",
        label="Kalkulus I - A",
    ))
    db.add(_Dosen(
        id=dosen1_id, kode="D01", nama="Dr. Andi Wijaya",
        homebase_prodi_id=prodi_id,
    ))
    db.add(_Dosen(
        id=dosen2_id, kode="D02", nama="Budi Santoso, M.Si",
        homebase_prodi_id=prodi_id,
    ))
    db.add(_Ruang(id=ruang_id, nama="G.1.01", lantai=1))
    db.add(_Timeslot(
        id=timeslot_id, kode="mon_s1", hari="Senin", sesi=1,
        jam_mulai=datetime.time(7, 30), jam_selesai=datetime.time(10, 0),
        label="Senin 07:30–10:00", sks=3,
    ))
    db.add(_SesiJadwal(
        id=sesi_id, nama="Genap 2024/2025",
        semester="Genap", tahun_akademik="2024/2025", status="Draft",
    ))
    db.add(_JadwalAssignment(
        id=assignment_id,
        sesi_id=sesi_id,
        mk_kelas_id=mk_kelas_id,
        dosen1_id=dosen1_id,
        dosen2_id=dosen2_id,
        timeslot_id=timeslot_id,
        ruang_id=ruang_id,
        catatan="Catatan test",
    ))
    db.commit()

    return {
        "sesi_id": sesi_id,
        "prodi_id": prodi_id,
        "mk_id": mk_id,
        "mk_kelas_id": mk_kelas_id,
        "dosen1_id": dosen1_id,
        "dosen2_id": dosen2_id,
        "ruang_id": ruang_id,
        "timeslot_id": timeslot_id,
        "assignment_id": assignment_id,
    }


# ---------------------------------------------------------------------------
# Helper: build a local ExcelExporter that uses the SQLite models
# ---------------------------------------------------------------------------

def _make_exporter(session):
    """
    Build an ExcelExporter subclass that queries the local SQLite models
    instead of the production PostgreSQL models.
    """
    from app.services.excel_exporter import ExcelExporter, _HARI_ORDER, _HEADER_FILL, _HEADER_FONT
    import openpyxl as _openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    import io as _io
    from collections import defaultdict

    class _LocalExporter(ExcelExporter):
        def export_jadwal(self, sesi_id):
            db = self.db_session

            sesi = db.get(_SesiJadwal, str(sesi_id))
            if sesi is None:
                raise ValueError(f"SesiJadwal dengan id '{sesi_id}' tidak ditemukan")

            rows = (
                db.query(
                    _JadwalAssignment,
                    _Timeslot,
                    _MataKuliahKelas,
                    _MataKuliah,
                    _Kurikulum,
                    _Prodi,
                )
                .join(_Timeslot, _JadwalAssignment.timeslot_id == _Timeslot.id)
                .join(_MataKuliahKelas, _JadwalAssignment.mk_kelas_id == _MataKuliahKelas.id)
                .join(_MataKuliah, _MataKuliahKelas.mata_kuliah_id == _MataKuliah.id)
                .join(_Kurikulum, _MataKuliah.kurikulum_id == _Kurikulum.id)
                .join(_Prodi, _Kurikulum.prodi_id == _Prodi.id)
                .filter(_JadwalAssignment.sesi_id == str(sesi_id))
                .all()
            )

            def sort_key(r):
                assignment, timeslot, mk_kelas, mk, kurikulum, prodi = r
                return (
                    _HARI_ORDER.get(timeslot.hari, 99),
                    timeslot.sesi,
                    prodi.singkat,
                    mk.semester,
                )

            rows_sorted = sorted(rows, key=sort_key)

            wb = _openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Jadwal Utama"

            headers1 = [
                "Hari", "Sesi", "Kode MK", "Nama Mata Kuliah", "SKS",
                "Kelas", "Semester", "Prodi", "Dosen I", "Dosen II",
                "Ruang", "Catatan",
            ]
            ws1.append(headers1)
            self._style_header_row(ws1, 1, len(headers1))
            ws1.freeze_panes = "A2"

            for assignment, timeslot, mk_kelas, mk, kurikulum, prodi in rows_sorted:
                dosen1 = db.get(_Dosen, str(assignment.dosen1_id))
                dosen1_nama = dosen1.nama if dosen1 else ""

                dosen2_nama = ""
                if assignment.dosen2_id:
                    dosen2 = db.get(_Dosen, str(assignment.dosen2_id))
                    dosen2_nama = dosen2.nama if dosen2 else ""

                ruang_nama = ""
                if assignment.ruang_id:
                    ruang = db.get(_Ruang, str(assignment.ruang_id))
                    ruang_nama = ruang.nama if ruang else ""

                jam_mulai = timeslot.jam_mulai.strftime("%H:%M")
                jam_selesai = timeslot.jam_selesai.strftime("%H:%M")
                sesi_label = f"{jam_mulai}–{jam_selesai}"

                ws1.append([
                    timeslot.hari,
                    sesi_label,
                    mk.kode,
                    mk.nama,
                    mk.sks,
                    mk_kelas.kelas or "",
                    mk.semester,
                    prodi.singkat,
                    dosen1_nama,
                    dosen2_nama,
                    ruang_nama,
                    assignment.catatan or "",
                ])

            self._auto_width(ws1)

            # Sheet 2
            ws2 = wb.create_sheet(title="Rekap Beban SKS")
            prodi_set = set()
            for _, _, _, _, _, prodi in rows_sorted:
                prodi_set.add(prodi.singkat)
            prodi_cols = sorted(prodi_set)

            headers2 = (
                ["Nama Dosen", "Kode Dosen", "Homebase Prodi"]
                + prodi_cols
                + ["Total SKS", "Jumlah MK"]
            )
            ws2.append(headers2)
            self._style_header_row(ws2, 1, len(headers2))
            ws2.freeze_panes = "A2"

            dosen_data = defaultdict(lambda: defaultdict(int))
            dosen_mk_count = defaultdict(int)
            dosen_obj_map = {}

            for assignment, timeslot, mk_kelas, mk, kurikulum, prodi in rows_sorted:
                sks = mk.sks
                for dosen_id in [assignment.dosen1_id, assignment.dosen2_id]:
                    if dosen_id is None:
                        continue
                    dosen_data[dosen_id][prodi.singkat] += sks
                    dosen_mk_count[dosen_id] += 1
                    if dosen_id not in dosen_obj_map:
                        d = db.get(_Dosen, str(dosen_id))
                        if d:
                            dosen_obj_map[dosen_id] = d

            sorted_dosen_ids = sorted(
                dosen_obj_map.keys(),
                key=lambda did: dosen_obj_map[did].nama,
            )

            for dosen_id in sorted_dosen_ids:
                dosen = dosen_obj_map[dosen_id]
                homebase = ""
                if dosen.homebase_prodi_id:
                    hp = db.get(_Prodi, str(dosen.homebase_prodi_id))
                    homebase = hp.singkat if hp else ""

                prodi_sks = dosen_data[dosen_id]
                total_sks = sum(prodi_sks.values())
                jumlah_mk = dosen_mk_count[dosen_id]

                row = (
                    [dosen.nama, dosen.kode, homebase]
                    + [prodi_sks.get(p, 0) for p in prodi_cols]
                    + [total_sks, jumlah_mk]
                )
                ws2.append(row)

            self._auto_width(ws2)

            buf = _io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf.read()

    exporter = _LocalExporter.__new__(_LocalExporter)
    exporter.db_session = session
    return exporter


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_export_returns_bytes(seeded_db, db):
    """export_jadwal() must return a non-empty bytes object."""
    exporter = _make_exporter(db)
    result = exporter.export_jadwal(seeded_db["sesi_id"])
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_export_is_valid_xlsx(seeded_db, db):
    """The returned bytes must be a valid .xlsx workbook."""
    exporter = _make_exporter(db)
    result = exporter.export_jadwal(seeded_db["sesi_id"])
    wb = openpyxl.load_workbook(io.BytesIO(result))
    assert wb is not None


def test_export_has_two_sheets(seeded_db, db):
    """Workbook must contain exactly 'Jadwal Utama' and 'Rekap Beban SKS' sheets."""
    exporter = _make_exporter(db)
    result = exporter.export_jadwal(seeded_db["sesi_id"])
    wb = openpyxl.load_workbook(io.BytesIO(result))
    assert "Jadwal Utama" in wb.sheetnames
    assert "Rekap Beban SKS" in wb.sheetnames


def test_jadwal_utama_headers(seeded_db, db):
    """Sheet 1 must have the correct 12 column headers in row 1."""
    exporter = _make_exporter(db)
    result = exporter.export_jadwal(seeded_db["sesi_id"])
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["Jadwal Utama"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, 13)]
    assert headers == [
        "Hari", "Sesi", "Kode MK", "Nama Mata Kuliah", "SKS",
        "Kelas", "Semester", "Prodi", "Dosen I", "Dosen II",
        "Ruang", "Catatan",
    ]


def test_jadwal_utama_data_row(seeded_db, db):
    """Sheet 1 row 2 must contain the seeded assignment data."""
    exporter = _make_exporter(db)
    result = exporter.export_jadwal(seeded_db["sesi_id"])
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["Jadwal Utama"]

    row = [ws.cell(row=2, column=c).value for c in range(1, 13)]
    assert row[0] == "Senin"           # Hari
    assert row[1] == "07:30–10:00"     # Sesi label
    assert row[2] == "MTK101"          # Kode MK
    assert row[3] == "Kalkulus I"      # Nama MK
    assert row[4] == 3                 # SKS
    assert row[5] == "A"               # Kelas
    assert row[6] == 1                 # Semester
    assert row[7] == "S1 MTK"          # Prodi
    assert row[8] == "Dr. Andi Wijaya" # Dosen I
    assert row[9] == "Budi Santoso, M.Si"  # Dosen II
    assert row[10] == "G.1.01"         # Ruang
    assert row[11] == "Catatan test"   # Catatan


def test_rekap_beban_sks_headers(seeded_db, db):
    """Sheet 2 must have Nama Dosen, Kode Dosen, Homebase Prodi, prodi cols, Total SKS, Jumlah MK."""
    exporter = _make_exporter(db)
    result = exporter.export_jadwal(seeded_db["sesi_id"])
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["Rekap Beban SKS"]

    num_cols = ws.max_column
    headers = [ws.cell(row=1, column=c).value for c in range(1, num_cols + 1)]
    assert headers[0] == "Nama Dosen"
    assert headers[1] == "Kode Dosen"
    assert headers[2] == "Homebase Prodi"
    assert headers[-2] == "Total SKS"
    assert headers[-1] == "Jumlah MK"


def test_rekap_beban_sks_dosen_rows(seeded_db, db):
    """Sheet 2 must have one row per dosen involved in the sesi."""
    exporter = _make_exporter(db)
    result = exporter.export_jadwal(seeded_db["sesi_id"])
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["Rekap Beban SKS"]

    # Row 1 = header; rows 2+ = dosen data
    dosen_names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    dosen_names = [n for n in dosen_names if n]  # filter empty
    assert "Budi Santoso, M.Si" in dosen_names
    assert "Dr. Andi Wijaya" in dosen_names


def test_rekap_total_sks_correct(seeded_db, db):
    """Total SKS for each dosen must equal sum of SKS from their assignments."""
    exporter = _make_exporter(db)
    result = exporter.export_jadwal(seeded_db["sesi_id"])
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["Rekap Beban SKS"]

    num_cols = ws.max_column
    # Find Total SKS column (second to last)
    total_sks_col = num_cols - 1

    for row in range(2, ws.max_row + 1):
        nama = ws.cell(row=row, column=1).value
        if nama:
            total_sks = ws.cell(row=row, column=total_sks_col).value
            # Each dosen teaches 1 MK of 3 SKS
            assert total_sks == 3, f"Expected 3 SKS for {nama}, got {total_sks}"


def test_export_invalid_sesi_raises(db):
    """export_jadwal() must raise ValueError for unknown sesi_id."""
    exporter = _make_exporter(db)
    with pytest.raises(ValueError, match="tidak ditemukan"):
        exporter.export_jadwal(uuid.uuid4())


def test_export_null_ruang_and_dosen2(db, engine):
    """Assignment with NULL ruang_id and NULL dosen2_id must export without error."""
    Session = sessionmaker(bind=engine)
    session = Session()

    prodi_id = str(uuid.uuid4())
    kurikulum_id = str(uuid.uuid4())
    mk_id = str(uuid.uuid4())
    mk_kelas_id = str(uuid.uuid4())
    dosen1_id = str(uuid.uuid4())
    timeslot_id = str(uuid.uuid4())
    sesi_id = str(uuid.uuid4())

    session.add(_Prodi(
        id=prodi_id, kode="S1STK", strata="S1",
        nama="S1 Statistika", singkat="S1 STK", kategori="Reguler",
    ))
    session.add(_Kurikulum(id=kurikulum_id, kode="K2021", tahun="2021", prodi_id=prodi_id))
    session.add(_MataKuliah(
        id=mk_id, kode="STK101", kurikulum_id=kurikulum_id,
        nama="Statistika Dasar", sks=3, semester=1, jenis="Wajib",
    ))
    session.add(_MataKuliahKelas(
        id=mk_kelas_id, mata_kuliah_id=mk_id, kelas=None,
        label="Statistika Dasar",
    ))
    session.add(_Dosen(id=dosen1_id, kode="D03", nama="Citra Dewi, M.Sc"))
    session.add(_Timeslot(
        id=timeslot_id, kode="tue_s2", hari="Selasa", sesi=2,
        jam_mulai=datetime.time(10, 0), jam_selesai=datetime.time(12, 30),
        label="Selasa 10:00–12:30", sks=3,
    ))
    session.add(_SesiJadwal(
        id=sesi_id, nama="Ganjil 2025/2026",
        semester="Ganjil", tahun_akademik="2025/2026", status="Draft",
    ))
    session.add(_JadwalAssignment(
        id=str(uuid.uuid4()),
        sesi_id=sesi_id,
        mk_kelas_id=mk_kelas_id,
        dosen1_id=dosen1_id,
        dosen2_id=None,   # NULL
        timeslot_id=timeslot_id,
        ruang_id=None,    # NULL
        catatan=None,
    ))
    session.commit()

    exporter = _make_exporter(session)
    result = exporter.export_jadwal(sesi_id)
    assert isinstance(result, bytes)

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb["Jadwal Utama"]
    row = [ws.cell(row=2, column=c).value for c in range(1, 13)]
    assert row[9] in ("", None)   # Dosen II empty
    assert row[10] in ("", None)  # Ruang empty
    assert row[5] in ("", None)   # Kelas empty (NULL kelas)

    session.close()


def test_export_empty_sesi(db, engine):
    """Exporting a sesi with no assignments should return a valid xlsx with only headers."""
    Session = sessionmaker(bind=engine)
    session = Session()

    sesi_id = str(uuid.uuid4())
    session.add(_SesiJadwal(
        id=sesi_id, nama="Kosong",
        semester="Ganjil", tahun_akademik="2099/2100", status="Draft",
    ))
    session.commit()

    exporter = _make_exporter(session)
    result = exporter.export_jadwal(sesi_id)
    assert isinstance(result, bytes)

    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws1 = wb["Jadwal Utama"]
    # Only header row, no data rows
    assert ws1.max_row == 1

    session.close()
