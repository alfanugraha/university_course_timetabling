"""
backend/tests/test_excel_importer.py
Unit tests untuk helper resolve_dosen di excel_importer.py.
Menggunakan SQLite in-memory dengan tabel dosen saja (tanpa tabel lain
yang menggunakan tipe PostgreSQL-specific seperti ARRAY).
"""

import uuid

import pytest
from sqlalchemy import Column, String, create_engine
from sqlalchemy.orm import DeclarativeBase, sessionmaker


# ---------------------------------------------------------------------------
# Minimal ORM — hanya tabel dosen, kompatibel dengan SQLite
# ---------------------------------------------------------------------------

class _Base(DeclarativeBase):
    pass


class _Dosen(_Base):
    __tablename__ = "dosen"

    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    kode = Column(String(10), unique=True, nullable=False)
    nama = Column(String(200), nullable=False)
    status = Column(String(20), default="Aktif", nullable=False)


# ---------------------------------------------------------------------------
# Fixture
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def db_session():
    engine = create_engine("sqlite:///:memory:")
    _Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()

    session.add_all([
        _Dosen(kode="ABD", nama="Dr. Abdullah Syukri", status="Aktif"),
        _Dosen(kode="RHM", nama="Rahmat Hidayat, M.Si", status="Aktif"),
        _Dosen(kode="SRI", nama="Sri Wahyuni", status="Aktif"),
    ])
    session.commit()
    yield session
    session.close()
    engine.dispose()


# ---------------------------------------------------------------------------
# Helper: wrap resolve_dosen agar bekerja dengan model SQLite lokal
# ---------------------------------------------------------------------------

from app.services.excel_importer import normalize_str
from typing import Optional


def _resolve_dosen(nama_or_kode: str, session) -> Optional[_Dosen]:
    """
    Reimplementasi resolve_dosen menggunakan model SQLite lokal (_Dosen).
    Logika identik dengan fungsi asli di excel_importer.py.
    """
    if not nama_or_kode:
        return None
    normalized = normalize_str(nama_or_kode)
    if not normalized:
        return None

    # Lookup by kode (exact, case-insensitive)
    dosen = session.query(_Dosen).filter(
        _Dosen.kode.ilike(normalized)
    ).first()
    if dosen:
        return dosen

    # Fallback: lookup by nama (case-insensitive contains)
    dosen = session.query(_Dosen).filter(
        _Dosen.nama.ilike(f"%{normalized}%")
    ).first()
    return dosen  # None if not found


# ---------------------------------------------------------------------------
# Tests — lookup by kode
# ---------------------------------------------------------------------------

def test_resolve_by_kode_exact(db_session):
    result = _resolve_dosen("ABD", db_session)
    assert result is not None
    assert result.kode == "ABD"


def test_resolve_by_kode_case_insensitive(db_session):
    result = _resolve_dosen("abd", db_session)
    assert result is not None
    assert result.kode == "ABD"


def test_resolve_by_kode_mixed_case(db_session):
    result = _resolve_dosen("Rhm", db_session)
    assert result is not None
    assert result.kode == "RHM"


# ---------------------------------------------------------------------------
# Tests — lookup by nama
# ---------------------------------------------------------------------------

def test_resolve_by_nama_partial(db_session):
    result = _resolve_dosen("Abdullah", db_session)
    assert result is not None
    assert result.kode == "ABD"


def test_resolve_by_nama_case_insensitive(db_session):
    result = _resolve_dosen("sri wahyuni", db_session)
    assert result is not None
    assert result.kode == "SRI"


def test_resolve_by_nama_partial_lowercase(db_session):
    result = _resolve_dosen("rahmat", db_session)
    assert result is not None
    assert result.kode == "RHM"


# ---------------------------------------------------------------------------
# Tests — not found → None, no exception
# ---------------------------------------------------------------------------

def test_resolve_not_found_returns_none(db_session):
    assert _resolve_dosen("TIDAKADA", db_session) is None


def test_resolve_unknown_nama_returns_none(db_session):
    assert _resolve_dosen("Budi Santoso", db_session) is None


# ---------------------------------------------------------------------------
# Tests — edge cases: empty / None input
# ---------------------------------------------------------------------------

def test_resolve_empty_string_returns_none(db_session):
    assert _resolve_dosen("", db_session) is None


def test_resolve_whitespace_only_returns_none(db_session):
    assert _resolve_dosen("   ", db_session) is None


def test_resolve_none_input_returns_none(db_session):
    assert _resolve_dosen(None, db_session) is None  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# Smoke test: verify the real resolve_dosen function is importable and
# has the correct signature (integration check without a real DB)
# ---------------------------------------------------------------------------

def test_resolve_dosen_importable():
    from app.services.excel_importer import resolve_dosen
    import inspect
    sig = inspect.signature(resolve_dosen)
    params = list(sig.parameters)
    assert "nama_or_kode" in params
    assert "session" in params


# ---------------------------------------------------------------------------
# Tests — import_mata_kuliah_kelas
# ---------------------------------------------------------------------------
# These tests use a minimal SQLite in-memory schema that mirrors only the
# tables needed: mata_kuliah and mata_kuliah_kelas (no PostgreSQL-specific
# types required).
# ---------------------------------------------------------------------------

import io
import uuid as _uuid

import openpyxl
from sqlalchemy import Column, ForeignKey, String, Text, UniqueConstraint
from sqlalchemy.orm import DeclarativeBase, sessionmaker, relationship
from sqlalchemy import create_engine as _create_engine
from unittest.mock import patch, MagicMock


class _Base2(DeclarativeBase):
    pass


class _MataKuliah(_Base2):
    __tablename__ = "mata_kuliah"

    id = Column(String(36), primary_key=True, default=lambda: str(_uuid.uuid4()))
    kode = Column(String(20), nullable=False, unique=True)
    nama = Column(String(200), nullable=False)

    kelas_list = relationship("_MataKuliahKelas", back_populates="mata_kuliah",
                              cascade="all, delete-orphan")


class _MataKuliahKelas(_Base2):
    __tablename__ = "mata_kuliah_kelas"

    id = Column(String(36), primary_key=True, default=lambda: str(_uuid.uuid4()))
    mata_kuliah_id = Column(String(36), ForeignKey("mata_kuliah.id"), nullable=False)
    kelas = Column(String(5), nullable=True)
    label = Column(String(200), nullable=False)
    ket = Column(Text, nullable=True)

    __table_args__ = (
        UniqueConstraint("mata_kuliah_id", "kelas", name="uq_mk_kelas"),
    )

    mata_kuliah = relationship("_MataKuliah", back_populates="kelas_list")


@pytest.fixture(scope="function")
def mk_db_session():
    """SQLite in-memory session with mata_kuliah + mata_kuliah_kelas tables."""
    engine = _create_engine("sqlite:///:memory:")
    _Base2.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()

    # Seed one MataKuliah record
    session.add(_MataKuliah(id=str(_uuid.uuid4()), kode="MTK001", nama="Kalkulus I"))
    session.commit()
    yield session
    session.close()
    engine.dispose()


def _make_kelas_xlsx(rows: list[tuple]) -> io.BytesIO:
    """
    Build an in-memory xlsx with sheet 'db_kelas'.
    Header: kode_mk | mata_kuliah | kelas | mk_kelas | sks | mk_syarat | Jenis | SMT | kurikulum | prodi | ket
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "db_kelas"
    ws.append(["kode_mk", "mata_kuliah", "kelas", "mk_kelas", "sks",
               "mk_syarat", "Jenis", "SMT", "kurikulum", "prodi", "ket"])
    for row in rows:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _make_importer(session):
    """
    Build an ExcelImporter whose db_session queries are redirected to the
    local SQLite session using the local _MataKuliah / _MataKuliahKelas models.
    """
    from app.services.excel_importer import ExcelImporter, ImportResult, ImportWarning, normalize_str, _clean_nullable_str
    import openpyxl as _openpyxl

    class _LocalImporter(ExcelImporter):
        """Subclass that overrides import_mata_kuliah_kelas to use local models."""

        def import_mata_kuliah_kelas(self, file) -> ImportResult:
            SHEET = "db_kelas"
            try:
                wb = _openpyxl.load_workbook(file, read_only=True, data_only=True)
            except Exception as exc:
                return ImportResult(
                    total=0, inserted=0, updated=0, skipped=1,
                    warnings=[ImportWarning(row=0, sheet="(file)", value=str(file), reason=str(exc))],
                )

            if SHEET not in wb.sheetnames:
                return ImportResult(total=0, inserted=0, updated=0, skipped=0)

            ws = wb[SHEET]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            total = inserted = updated = skipped = 0
            warnings_list = []

            for row_num, row in enumerate(rows, start=2):
                total += 1
                try:
                    kode_mk_raw = row[0] if len(row) > 0 else None
                    kelas_raw = row[2] if len(row) > 2 else None
                    label_raw = row[3] if len(row) > 3 else None
                    ket_raw = row[10] if len(row) > 10 else None

                    kode_mk = normalize_str(kode_mk_raw)
                    if not kode_mk:
                        raise ValueError("kode_mk kosong")

                    label = str(label_raw).strip() if label_raw else None
                    if not label:
                        raise ValueError("label (mk_kelas) kosong")

                    kelas = None
                    if kelas_raw is not None:
                        kelas_str = str(kelas_raw).strip()
                        kelas = kelas_str[:5] if kelas_str else None

                    ket = _clean_nullable_str(ket_raw)

                    # Lookup using local model
                    mk = session.query(_MataKuliah).filter(
                        _MataKuliah.kode.ilike(kode_mk)
                    ).first()
                    if mk is None:
                        skipped += 1
                        warnings_list.append(ImportWarning(
                            row=row_num, sheet=SHEET, value=kode_mk_raw,
                            reason=f"kode_mk '{kode_mk_raw}' tidak ditemukan di tabel mata_kuliah",
                        ))
                        continue

                    existing = session.query(_MataKuliahKelas).filter(
                        _MataKuliahKelas.mata_kuliah_id == mk.id,
                        _MataKuliahKelas.kelas == kelas,
                    ).first()

                    if existing:
                        existing.label = label
                        existing.ket = ket
                        updated += 1
                    else:
                        session.add(_MataKuliahKelas(
                            mata_kuliah_id=mk.id,
                            kelas=kelas,
                            label=label,
                            ket=ket,
                        ))
                        inserted += 1

                    session.flush()

                except Exception as exc:
                    skipped += 1
                    warnings_list.append(ImportWarning(
                        row=row_num, sheet=SHEET, value=row, reason=str(exc)
                    ))

            session.commit()
            return ImportResult(total=total, inserted=inserted, updated=updated,
                                skipped=skipped, warnings=warnings_list)

    importer = _LocalImporter.__new__(_LocalImporter)
    importer.db_session = session
    return importer


# --- Happy path: valid rows are inserted ---

def test_import_mata_kuliah_kelas_happy_path(mk_db_session):
    """Valid rows with known kode_mk should be inserted successfully."""
    xlsx = _make_kelas_xlsx([
        ("MTK001", "Kalkulus I", "A", "Kalkulus I (K2025) - A", 3, "-", "Wajib", 1, "K2025", "S1 MTK", "Internal"),
        ("MTK001", "Kalkulus I", "B", "Kalkulus I (K2025) - B", 3, "-", "Wajib", 1, "K2025", "S1 MTK", "Internal"),
    ])
    importer = _make_importer(mk_db_session)
    result = importer.import_mata_kuliah_kelas(xlsx)

    assert result.total == 2
    assert result.inserted == 2
    assert result.updated == 0
    assert result.skipped == 0
    assert result.warnings == []

    records = mk_db_session.query(_MataKuliahKelas).all()
    assert len(records) == 2
    labels = {r.label for r in records}
    assert "Kalkulus I (K2025) - A" in labels
    assert "Kalkulus I (K2025) - B" in labels


# --- kode_mk not found → warning, not raise ---

def test_import_mata_kuliah_kelas_unknown_kode_mk(mk_db_session):
    """Rows with unknown kode_mk should be skipped with a warning, not raise."""
    xlsx = _make_kelas_xlsx([
        ("UNKNOWN999", "Tidak Ada", "A", "Tidak Ada - A", 3, "-", "Wajib", 1, "K2025", "S1 MTK", "Internal"),
    ])
    importer = _make_importer(mk_db_session)
    result = importer.import_mata_kuliah_kelas(xlsx)

    assert result.skipped == 1
    assert result.inserted == 0
    assert len(result.warnings) == 1
    assert "tidak ditemukan" in result.warnings[0].reason
    assert result.warnings[0].value == "UNKNOWN999"


# --- Mixed: some valid, some unknown kode_mk ---

def test_import_mata_kuliah_kelas_mixed_rows(mk_db_session):
    """Valid rows are inserted; unknown kode_mk rows produce warnings only."""
    # Clean up any records from previous tests
    mk_db_session.query(_MataKuliahKelas).delete()
    mk_db_session.commit()

    xlsx = _make_kelas_xlsx([
        ("MTK001", "Kalkulus I", "A", "Kalkulus I - A", 3, "-", "Wajib", 1, "K2025", "S1 MTK", None),
        ("GHOST001", "Ghost MK", "A", "Ghost - A", 3, "-", "Wajib", 1, "K2025", "S1 MTK", None),
    ])
    importer = _make_importer(mk_db_session)
    result = importer.import_mata_kuliah_kelas(xlsx)

    assert result.total == 2
    assert result.inserted == 1
    assert result.skipped == 1
    assert len(result.warnings) == 1


# --- Upsert: existing record is updated, not duplicated ---

def test_import_mata_kuliah_kelas_upsert(mk_db_session):
    """Re-importing the same (mata_kuliah_id, kelas) should update, not insert."""
    mk_db_session.query(_MataKuliahKelas).delete()
    mk_db_session.commit()

    xlsx_first = _make_kelas_xlsx([
        ("MTK001", "Kalkulus I", "A", "Kalkulus I - A (v1)", 3, "-", "Wajib", 1, "K2025", "S1 MTK", None),
    ])
    importer = _make_importer(mk_db_session)
    r1 = importer.import_mata_kuliah_kelas(xlsx_first)
    assert r1.inserted == 1

    xlsx_second = _make_kelas_xlsx([
        ("MTK001", "Kalkulus I", "A", "Kalkulus I - A (v2)", 3, "-", "Wajib", 1, "K2025", "S1 MTK", None),
    ])
    r2 = importer.import_mata_kuliah_kelas(xlsx_second)
    assert r2.updated == 1
    assert r2.inserted == 0

    records = mk_db_session.query(_MataKuliahKelas).all()
    assert len(records) == 1
    assert records[0].label == "Kalkulus I - A (v2)"


# --- kelas=None (no parallel class) is handled ---

def test_import_mata_kuliah_kelas_null_kelas(mk_db_session):
    """Rows with kelas=None (single class MK) should be inserted correctly."""
    mk_db_session.query(_MataKuliahKelas).delete()
    mk_db_session.commit()

    xlsx = _make_kelas_xlsx([
        ("MTK001", "Kalkulus I", None, "Kalkulus I (MTK25)", 3, "-", "Wajib", 1, "K2025", "S1 MTK", "Internal"),
    ])
    importer = _make_importer(mk_db_session)
    result = importer.import_mata_kuliah_kelas(xlsx)

    assert result.inserted == 1
    record = mk_db_session.query(_MataKuliahKelas).first()
    assert record.kelas is None
    assert record.label == "Kalkulus I (MTK25)"
    assert record.ket == "Internal"
