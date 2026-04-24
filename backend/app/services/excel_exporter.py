"""
backend/app/services/excel_exporter.py
ExcelExporter — Fase 5 (Export Excel)

Sistem Penjadwalan Kuliah, Jurusan Matematika FMIPA UNRI

Modul ini berisi ExcelExporter class untuk mengekspor data jadwal
dari database ke format file Excel (.xlsx) sesuai templat standar jurusan.
"""

from __future__ import annotations

import io
import logging
from collections import defaultdict
from uuid import UUID

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.dosen import Dosen
from app.models.jadwal_assignment import JadwalAssignment
from app.models.kurikulum import Kurikulum
from app.models.mata_kuliah import MataKuliah, MataKuliahKelas
from app.models.prodi import Prodi
from app.models.ruang import Ruang
from app.models.sesi_jadwal import SesiJadwal
from app.models.timeslot import Timeslot

logger = logging.getLogger(__name__)

# Urutan hari untuk sorting
_HARI_ORDER = {
    "Senin": 1,
    "Selasa": 2,
    "Rabu": 3,
    "Kamis": 4,
    "Jumat": 5,
}

# Warna header
_HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_HEADER_FONT = Font(bold=True)


class ExcelExporter:
    """
    Service untuk mengekspor data jadwal kuliah ke file Excel.

    Menghasilkan file .xlsx berisi jadwal lengkap untuk satu sesi penjadwalan,
    dengan format sesuai templat standar Jurusan Matematika FMIPA UNRI.

    Usage:
        exporter = ExcelExporter(db_session=session)
        excel_bytes = exporter.export_jadwal(sesi_id=some_uuid)
        # Kembalikan excel_bytes sebagai response file download
    """

    def __init__(self, db_session: Session) -> None:
        """
        Inisialisasi ExcelExporter dengan SQLAlchemy session.

        Args:
            db_session: SQLAlchemy Session yang aktif untuk query data jadwal.
        """
        self.db_session = db_session

    def export_jadwal(self, sesi_id: UUID) -> bytes:
        """
        Ekspor seluruh jadwal assignment untuk satu sesi ke format Excel (.xlsx).

        Mengambil semua JadwalAssignment yang terkait dengan sesi_id,
        lalu menyusunnya ke dalam workbook Excel dengan format standar jurusan.

        Args:
            sesi_id: UUID dari SesiJadwal yang akan diekspor.

        Returns:
            bytes: Konten file Excel (.xlsx) sebagai byte string,
                   siap dikirim sebagai HTTP response (Content-Type:
                   application/vnd.openxmlformats-officedocument.spreadsheetml.sheet).

        Raises:
            ValueError: Jika sesi_id tidak ditemukan di database.
        """
        db = self.db_session

        # Validate sesi exists
        sesi = db.get(SesiJadwal, sesi_id)
        if sesi is None:
            raise ValueError(f"SesiJadwal dengan id '{sesi_id}' tidak ditemukan")

        # Fetch all assignments with related data via joins
        rows = (
            db.query(
                JadwalAssignment,
                Timeslot,
                MataKuliahKelas,
                MataKuliah,
                Kurikulum,
                Prodi,
            )
            .join(Timeslot, JadwalAssignment.timeslot_id == Timeslot.id)
            .join(MataKuliahKelas, JadwalAssignment.mk_kelas_id == MataKuliahKelas.id)
            .join(MataKuliah, MataKuliahKelas.mata_kuliah_id == MataKuliah.id)
            .join(Kurikulum, MataKuliah.kurikulum_id == Kurikulum.id)
            .join(Prodi, Kurikulum.prodi_id == Prodi.id)
            .filter(JadwalAssignment.sesi_id == sesi_id)
            .all()
        )

        # Sort: Hari order, then sesi number, then prodi singkat, then semester
        def sort_key(r):
            assignment, timeslot, mk_kelas, mk, kurikulum, prodi = r
            return (
                _HARI_ORDER.get(timeslot.hari, 99),
                timeslot.sesi,
                prodi.singkat,
                mk.semester,
            )

        rows_sorted = sorted(rows, key=sort_key)

        wb = openpyxl.Workbook()

        # ------------------------------------------------------------------ #
        # Sheet 1 — Jadwal Utama
        # ------------------------------------------------------------------ #
        ws1 = wb.active
        ws1.title = "Jadwal Utama"

        headers1 = [
            "Hari",
            "Sesi",
            "Kode MK",
            "Nama Mata Kuliah",
            "SKS",
            "Kelas",
            "Semester",
            "Prodi",
            "Dosen I",
            "Dosen II",
            "Ruang",
            "Catatan",
        ]
        ws1.append(headers1)
        self._style_header_row(ws1, 1, len(headers1))
        ws1.freeze_panes = "A2"

        for assignment, timeslot, mk_kelas, mk, kurikulum, prodi in rows_sorted:
            # Resolve dosen names
            dosen1 = db.get(Dosen, assignment.dosen1_id)
            dosen1_nama = dosen1.nama if dosen1 else ""

            dosen2_nama = ""
            if assignment.dosen2_id:
                dosen2 = db.get(Dosen, assignment.dosen2_id)
                dosen2_nama = dosen2.nama if dosen2 else ""

            # Resolve ruang name
            ruang_nama = ""
            if assignment.ruang_id:
                ruang = db.get(Ruang, assignment.ruang_id)
                ruang_nama = ruang.nama if ruang else ""

            # Sesi label: "Sesi N (HH:MM–HH:MM)"
            jam_mulai = timeslot.jam_mulai.strftime("%H:%M")
            jam_selesai = timeslot.jam_selesai.strftime("%H:%M")
            sesi_label = f"{jam_mulai}–{jam_selesai}"

            ws1.append([
                timeslot.hari,
                sesi_label,
                mk.kode,
                mk.nama,
                mk.sks,
                mk_kelas.kelas or "",
                mk.semester,
                prodi.singkat,
                dosen1_nama,
                dosen2_nama,
                ruang_nama,
                assignment.catatan or "",
            ])

        self._auto_width(ws1)

        # ------------------------------------------------------------------ #
        # Sheet 2 — Rekap Beban SKS
        # ------------------------------------------------------------------ #
        ws2 = wb.create_sheet(title="Rekap Beban SKS")

        # Collect all prodi singkat values present in this sesi (sorted)
        prodi_set: set[str] = set()
        for _, _, _, _, _, prodi in rows_sorted:
            prodi_set.add(prodi.singkat)
        prodi_cols = sorted(prodi_set)

        headers2 = (
            ["Nama Dosen", "Kode Dosen", "Homebase Prodi"]
            + prodi_cols
            + ["Total SKS", "Jumlah MK"]
        )
        ws2.append(headers2)
        self._style_header_row(ws2, 1, len(headers2))
        ws2.freeze_panes = "A2"

        # Aggregate per dosen
        # dosen_data[dosen_id] = {prodi_singkat: sks_total, ...}
        dosen_data: dict[UUID, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        dosen_mk_count: dict[UUID, int] = defaultdict(int)
        dosen_obj_map: dict[UUID, Dosen] = {}

        for assignment, timeslot, mk_kelas, mk, kurikulum, prodi in rows_sorted:
            sks = mk.sks

            for dosen_id in [assignment.dosen1_id, assignment.dosen2_id]:
                if dosen_id is None:
                    continue
                dosen_data[dosen_id][prodi.singkat] += sks
                dosen_mk_count[dosen_id] += 1
                if dosen_id not in dosen_obj_map:
                    d = db.get(Dosen, dosen_id)
                    if d:
                        dosen_obj_map[dosen_id] = d

        # Sort dosen by name
        sorted_dosen_ids = sorted(
            dosen_obj_map.keys(),
            key=lambda did: dosen_obj_map[did].nama,
        )

        for dosen_id in sorted_dosen_ids:
            dosen = dosen_obj_map[dosen_id]

            # Homebase prodi singkat
            homebase = ""
            if dosen.homebase_prodi_id:
                hp = db.get(Prodi, dosen.homebase_prodi_id)
                homebase = hp.singkat if hp else ""

            prodi_sks = dosen_data[dosen_id]
            total_sks = sum(prodi_sks.values())
            jumlah_mk = dosen_mk_count[dosen_id]

            row = (
                [dosen.nama, dosen.kode, homebase]
                + [prodi_sks.get(p, 0) for p in prodi_cols]
                + [total_sks, jumlah_mk]
            )
            ws2.append(row)

        self._auto_width(ws2)

        # ------------------------------------------------------------------ #
        # Serialize to bytes
        # ------------------------------------------------------------------ #
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    # ---------------------------------------------------------------------- #
    # Helpers
    # ---------------------------------------------------------------------- #

    def _style_header_row(self, ws, row: int, num_cols: int) -> None:
        """Apply bold + light-blue fill to the header row."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _auto_width(self, ws) -> None:
        """Set approximate column widths based on max content length."""
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            # Clamp between 8 and 50 characters
            ws.column_dimensions[col_letter].width = max(8, min(max_len + 2, 50))
