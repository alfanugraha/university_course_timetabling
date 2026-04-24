"""
Excel Importer — Fase 5
Sistem Penjadwalan Kuliah, Jurusan Matematika FMIPA UNRI

Modul ini berisi dataclass untuk hasil import Excel.
Business logic (ExcelImporter class) akan ditambahkan di T5.1.1+.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Optional

import openpyxl
from sqlalchemy.orm import Session

from app.models.dosen import Dosen
from app.models.jadwal_assignment import JadwalAssignment
from app.models.kurikulum import Kurikulum
from app.models.mata_kuliah import MataKuliah, MataKuliahKelas
from app.models.prodi import Prodi
from app.models.ruang import Ruang
from app.models.timeslot import Timeslot

logger = logging.getLogger(__name__)


def normalize_str(val) -> str:
    """
    Normalisasi nilai string dari sel Excel untuk keperluan lookup.

    - None atau non-string → kembalikan string kosong ""
    - String → strip whitespace leading/trailing, lalu lowercase

    Tidak pernah raise exception (tolerant import strategy).
    """
    if val is None:
        return ""
    try:
        return str(val).strip().lower()
    except Exception:
        return ""


def resolve_dosen(nama_or_kode: str, session: Session) -> Optional[Dosen]:
    """
    Lookup dosen dari database berdasarkan kode atau nama.

    Urutan lookup:
    1. Coba cocokkan dengan kolom `kode` (exact match, case-insensitive via normalize_str)
    2. Jika tidak ditemukan, coba cocokkan dengan kolom `nama` (case-insensitive contains)

    Mengembalikan None (tidak raise exception) jika:
    - Input kosong / None
    - Tidak ada dosen yang cocok

    Caller bertanggung jawab mencatat warning jika None dikembalikan.
    """
    if not nama_or_kode:
        return None

    normalized = normalize_str(nama_or_kode)
    if not normalized:
        return None

    # Coba match by kode (exact, case-insensitive)
    dosen = session.query(Dosen).filter(
        Dosen.kode.ilike(normalized)
    ).first()
    if dosen:
        return dosen

    # Fallback: match by nama (case-insensitive contains)
    dosen = session.query(Dosen).filter(
        Dosen.nama.ilike(f"%{normalized}%")
    ).first()
    if dosen:
        return dosen

    logger.warning("resolve_dosen: dosen tidak ditemukan untuk %r", nama_or_kode)
    return None


def _clean_nullable_str(val, max_len: int = 255) -> Optional[str]:
    """
    Bersihkan nilai string yang boleh NULL dari sel Excel.

    - None atau string kosong setelah strip → kembalikan None
    - String valid → strip whitespace, potong ke max_len

    Tidak pernah raise exception.
    """
    if val is None:
        return None
    try:
        s = str(val).strip()
        return s[:max_len] if s else None
    except Exception:
        return None


@dataclass
class ImportWarning:
    """
    Catatan satu baris bermasalah selama proses import Excel.

    Fields:
        row     : Nomor baris di sheet Excel yang bermasalah (1-indexed)
        sheet   : Nama sheet Excel tempat baris berada
        value   : Nilai bermasalah (bisa berupa string, angka, atau None)
        reason  : Alasan baris dilewati / tidak dapat diproses
    """

    row: int
    sheet: str
    value: Any
    reason: str


@dataclass
class ImportResult:
    """
    Ringkasan hasil proses import Excel.

    Fields:
        total    : Total baris yang diproses
        inserted : Jumlah baris berhasil di-insert (record baru)
        updated  : Jumlah baris berhasil di-update (record sudah ada)
        skipped  : Jumlah baris yang dilewati karena bermasalah
        warnings : Daftar detail baris yang dilewati
    """

    total: int
    inserted: int
    updated: int
    skipped: int
    warnings: list[ImportWarning] = field(default_factory=list)


class ExcelImporter:
    """
    Service untuk mengimpor data dari file Excel ke database.

    Menggunakan strategi tolerant import — satu baris gagal tidak membatalkan
    seluruh proses import. Setiap baris bermasalah dicatat sebagai ImportWarning.

    Usage:
        importer = ExcelImporter(db_session=session)
        result = importer.import_master_db(file)
    """

    def __init__(self, db_session: Session) -> None:
        """
        Inisialisasi ExcelImporter dengan SQLAlchemy session.

        Args:
            db_session: SQLAlchemy Session yang aktif untuk operasi DB.
        """
        self.db_session = db_session

    def import_master_db(self, file) -> ImportResult:
        """
        Import data master dari file db.xlsx.

        Sheet yang diproses (urutan penting — FK dependency):
          1. Prodi         (sheet: 'Prodi')
          2. Kurikulum     (sheet: 'Kurikulum')
          3. Mata Kuliah   (sheet: 'Mata Kuliah')
          4. Ruang         (sheet: 'Ruang Kuliah')
          5. Dosen         (sheet: 'Dosen')

        Setiap sheet diproses secara tolerant — baris bermasalah dilewati
        dan dicatat sebagai warning tanpa membatalkan import keseluruhan.

        Args:
            file: File-like object atau path ke db.xlsx.

        Returns:
            ImportResult dengan ringkasan total, inserted, updated, skipped,
            dan daftar warnings per baris bermasalah.
        """
        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception as exc:
            logger.error("import_master_db: gagal membuka file Excel: %s", exc)
            return ImportResult(
                total=0, inserted=0, updated=0, skipped=1,
                warnings=[ImportWarning(row=0, sheet="(file)", value=str(file), reason=str(exc))],
            )

        total = inserted = updated = skipped = 0
        warnings: list[ImportWarning] = []

        def _merge(r: ImportResult) -> None:
            nonlocal total, inserted, updated, skipped
            total += r.total
            inserted += r.inserted
            updated += r.updated
            skipped += r.skipped
            warnings.extend(r.warnings)

        _merge(self._import_prodi(wb))
        _merge(self._import_kurikulum(wb))
        _merge(self._import_mata_kuliah(wb))
        _merge(self._import_ruang(wb))
        _merge(self._import_dosen(wb))

        return ImportResult(total=total, inserted=inserted, updated=updated,
                            skipped=skipped, warnings=warnings)

    # ------------------------------------------------------------------
    # Private helpers — satu method per sheet
    # ------------------------------------------------------------------

    def _import_prodi(self, wb: openpyxl.Workbook) -> ImportResult:
        """
        Import sheet 'Prodi'.

        Kolom Excel: id_prodi | strata | nama_prodi | kategori

        Karena sheet tidak memiliki kolom `kode` eksplisit, kode di-generate
        dari kombinasi strata + nama (misal: S-1 Matematika → S1MATH).
        Upsert berdasarkan kode yang di-generate.

        Mapping singkat:
          strata    → strata
          nama_prodi → nama
          kategori  → kategori
          kode      → di-generate: "{strata_clean}{nama_abbr}"
          singkat   → "{strata} {nama_abbr}"
        """
        SHEET = "Prodi"
        if SHEET not in wb.sheetnames:
            logger.warning("_import_prodi: sheet '%s' tidak ditemukan", SHEET)
            return ImportResult(total=0, inserted=0, updated=0, skipped=0)

        ws = wb[SHEET]
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
        total = inserted = updated = skipped = 0
        warnings: list[ImportWarning] = []

        for row_num, row in enumerate(rows, start=2):
            total += 1
            try:
                _, strata_raw, nama_raw, kategori_raw = row[:4]

                strata = str(strata_raw).strip() if strata_raw else None
                nama = str(nama_raw).strip() if nama_raw else None
                kategori = str(kategori_raw).strip() if kategori_raw else None

                if not strata or not nama:
                    raise ValueError("strata atau nama_prodi kosong")

                # Generate kode unik: "S1MATH", "S2MATH", "S1STAT", dll.
                strata_clean = strata.replace("-", "").replace(" ", "").upper()  # S1, S2
                # Ambil 4 huruf pertama nama sebagai abbreviasi
                nama_abbr = "".join(w[:4].upper() for w in nama.split()[:1])
                kode = f"{strata_clean}{nama_abbr}"[:10]
                singkat = f"{strata} {nama_abbr}"[:20]

                if not kategori:
                    kategori = "Internal"

                existing = self.db_session.query(Prodi).filter(
                    Prodi.kode == kode
                ).first()

                if existing:
                    existing.strata = strata
                    existing.nama = nama
                    existing.singkat = singkat
                    existing.kategori = kategori
                    updated += 1
                else:
                    self.db_session.add(Prodi(
                        kode=kode,
                        strata=strata,
                        nama=nama,
                        singkat=singkat,
                        kategori=kategori,
                    ))
                    inserted += 1

                self.db_session.flush()

            except Exception as exc:
                skipped += 1
                warnings.append(ImportWarning(
                    row=row_num, sheet=SHEET, value=row, reason=str(exc)
                ))
                logger.warning("_import_prodi row %d: %s", row_num, exc)

        self.db_session.commit()
        return ImportResult(total=total, inserted=inserted, updated=updated,
                            skipped=skipped, warnings=warnings)

    def _import_kurikulum(self, wb: openpyxl.Workbook) -> ImportResult:
        """
        Import sheet 'Kurikulum'.

        Kolom Excel: kode_kurikulum | kurikulum (tahun) | prodi (nama prodi)

        Upsert berdasarkan kode_kurikulum.
        Lookup prodi berdasarkan nama (case-insensitive contains).
        """
        SHEET = "Kurikulum"
        if SHEET not in wb.sheetnames:
            logger.warning("_import_kurikulum: sheet '%s' tidak ditemukan", SHEET)
            return ImportResult(total=0, inserted=0, updated=0, skipped=0)

        ws = wb[SHEET]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = inserted = updated = skipped = 0
        warnings: list[ImportWarning] = []

        for row_num, row in enumerate(rows, start=2):
            total += 1
            try:
                kode_raw, tahun_raw, prodi_nama_raw = row[:3]

                kode = normalize_str(kode_raw)
                if not kode:
                    raise ValueError("kode_kurikulum kosong")

                tahun = str(int(tahun_raw)) if tahun_raw else None
                if not tahun:
                    raise ValueError("tahun kurikulum kosong")

                prodi_nama = normalize_str(prodi_nama_raw)
                if not prodi_nama:
                    raise ValueError("nama prodi kosong")

                # Lookup prodi — tolerant: skip jika tidak ditemukan
                prodi = self.db_session.query(Prodi).filter(
                    Prodi.nama.ilike(f"%{prodi_nama}%")
                ).first()
                if prodi is None:
                    raise ValueError(f"prodi '{prodi_nama_raw}' tidak ditemukan di DB")

                existing = self.db_session.query(Kurikulum).filter(
                    Kurikulum.kode == kode.upper()
                ).first()

                if existing:
                    existing.tahun = tahun
                    existing.prodi_id = prodi.id
                    updated += 1
                else:
                    self.db_session.add(Kurikulum(
                        kode=kode.upper(),
                        tahun=tahun,
                        prodi_id=prodi.id,
                    ))
                    inserted += 1

                self.db_session.flush()

            except Exception as exc:
                skipped += 1
                warnings.append(ImportWarning(
                    row=row_num, sheet=SHEET, value=row, reason=str(exc)
                ))
                logger.warning("_import_kurikulum row %d: %s", row_num, exc)

        self.db_session.commit()
        return ImportResult(total=total, inserted=inserted, updated=updated,
                            skipped=skipped, warnings=warnings)

    def _import_mata_kuliah(self, wb: openpyxl.Workbook) -> ImportResult:
        """
        Import sheet 'Mata Kuliah'.

        Kolom Excel: kode_kurikulum | kode_mk | mata_kuliah | kategori |
                     smt | sifat | sks | (kosong) | prodi

        Upsert berdasarkan (kode_mk, kurikulum_id).
        Lookup kurikulum berdasarkan kode_kurikulum.
        """
        SHEET = "Mata Kuliah"
        if SHEET not in wb.sheetnames:
            logger.warning("_import_mata_kuliah: sheet '%s' tidak ditemukan", SHEET)
            return ImportResult(total=0, inserted=0, updated=0, skipped=0)

        ws = wb[SHEET]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = inserted = updated = skipped = 0
        warnings: list[ImportWarning] = []

        for row_num, row in enumerate(rows, start=2):
            total += 1
            try:
                kode_kurikulum_raw = row[0]
                kode_mk_raw = row[1]
                nama_raw = row[2]
                # row[3] = kategori (internal/layanan) — tidak dipakai di model MK
                smt_raw = row[4]
                sifat_raw = row[5]
                sks_raw = row[6]
                # row[7] = kosong
                # row[8] = prodi — tidak dipakai langsung (sudah via kurikulum)

                kode_kurikulum = normalize_str(kode_kurikulum_raw)
                kode_mk = normalize_str(kode_mk_raw)
                nama = str(nama_raw).strip() if nama_raw else None

                if not kode_kurikulum:
                    raise ValueError("kode_kurikulum kosong")
                if not kode_mk:
                    raise ValueError("kode_mk kosong")
                if not nama:
                    raise ValueError("nama mata kuliah kosong")

                try:
                    sks = int(sks_raw)
                except (TypeError, ValueError):
                    raise ValueError(f"sks tidak valid: {sks_raw!r}")

                try:
                    semester = int(smt_raw)
                except (TypeError, ValueError):
                    raise ValueError(f"semester tidak valid: {smt_raw!r}")

                # Normalise jenis: Wajib / Pilihan
                sifat_norm = normalize_str(sifat_raw)
                if "pilihan" in sifat_norm:
                    jenis = "Pilihan"
                else:
                    jenis = "Wajib"

                # Lookup kurikulum
                kurikulum = self.db_session.query(Kurikulum).filter(
                    Kurikulum.kode == kode_kurikulum.upper()
                ).first()
                if kurikulum is None:
                    raise ValueError(
                        f"kurikulum '{kode_kurikulum_raw}' tidak ditemukan di DB"
                    )

                existing = self.db_session.query(MataKuliah).filter(
                    MataKuliah.kode == kode_mk.upper(),
                    MataKuliah.kurikulum_id == kurikulum.id,
                ).first()

                if existing:
                    existing.nama = nama
                    existing.sks = sks
                    existing.semester = semester
                    existing.jenis = jenis
                    updated += 1
                else:
                    self.db_session.add(MataKuliah(
                        kode=kode_mk.upper(),
                        kurikulum_id=kurikulum.id,
                        nama=nama,
                        sks=sks,
                        semester=semester,
                        jenis=jenis,
                    ))
                    inserted += 1

                self.db_session.flush()

            except Exception as exc:
                skipped += 1
                warnings.append(ImportWarning(
                    row=row_num, sheet=SHEET, value=row, reason=str(exc)
                ))
                logger.warning("_import_mata_kuliah row %d: %s", row_num, exc)

        self.db_session.commit()
        return ImportResult(total=total, inserted=inserted, updated=updated,
                            skipped=skipped, warnings=warnings)

    def _import_ruang(self, wb: openpyxl.Workbook) -> ImportResult:
        """
        Import sheet 'Ruang Kuliah'.

        Kolom Excel: id_ruang (ignored) | nama_ruang | lantai | gedung

        Upsert berdasarkan nama_ruang (unique).
        """
        SHEET = "Ruang Kuliah"
        if SHEET not in wb.sheetnames:
            logger.warning("_import_ruang: sheet '%s' tidak ditemukan", SHEET)
            return ImportResult(total=0, inserted=0, updated=0, skipped=0)

        ws = wb[SHEET]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = inserted = updated = skipped = 0
        warnings: list[ImportWarning] = []

        for row_num, row in enumerate(rows, start=2):
            total += 1
            try:
                _, nama_raw, lantai_raw, gedung_raw = row[:4]

                nama = str(nama_raw).strip() if nama_raw else None
                if not nama:
                    raise ValueError("nama_ruang kosong")

                lantai: Optional[int] = None
                if lantai_raw is not None:
                    try:
                        lantai = int(lantai_raw)
                    except (TypeError, ValueError):
                        pass  # lantai tetap None jika tidak valid

                gedung = str(gedung_raw).strip() if gedung_raw else None

                existing = self.db_session.query(Ruang).filter(
                    Ruang.nama == nama
                ).first()

                if existing:
                    existing.lantai = lantai
                    existing.gedung = gedung
                    updated += 1
                else:
                    self.db_session.add(Ruang(
                        nama=nama,
                        lantai=lantai,
                        gedung=gedung,
                    ))
                    inserted += 1

                self.db_session.flush()

            except Exception as exc:
                skipped += 1
                warnings.append(ImportWarning(
                    row=row_num, sheet=SHEET, value=row, reason=str(exc)
                ))
                logger.warning("_import_ruang row %d: %s", row_num, exc)

        self.db_session.commit()
        return ImportResult(total=total, inserted=inserted, updated=updated,
                            skipped=skipped, warnings=warnings)

    def _import_dosen(self, wb: openpyxl.Workbook) -> ImportResult:
        """
        Import sheet 'Dosen'.

        Kolom Excel: nidn_nuptk | nip | kode | nama | jabfung | kjfd |
                     homebase | tgl_lahir | usia (ignored) | thn_angkat (ignored) | (extra)

        Upsert berdasarkan kode (wajib ada).
        nidn dan nip bersifat NULLABLE — toleran jika kosong atau duplikat.
        homebase di-resolve ke prodi via nama (case-insensitive contains).
        tgl_lahir di-parse dari datetime Excel atau string.
        """
        SHEET = "Dosen"
        if SHEET not in wb.sheetnames:
            logger.warning("_import_dosen: sheet '%s' tidak ditemukan", SHEET)
            return ImportResult(total=0, inserted=0, updated=0, skipped=0)

        ws = wb[SHEET]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = inserted = updated = skipped = 0
        warnings: list[ImportWarning] = []

        for row_num, row in enumerate(rows, start=2):
            total += 1
            try:
                nidn_raw = row[0] if len(row) > 0 else None
                nip_raw = row[1] if len(row) > 1 else None
                kode_raw = row[2] if len(row) > 2 else None
                nama_raw = row[3] if len(row) > 3 else None
                jabfung_raw = row[4] if len(row) > 4 else None
                kjfd_raw = row[5] if len(row) > 5 else None
                homebase_raw = row[6] if len(row) > 6 else None
                tgl_lahir_raw = row[7] if len(row) > 7 else None

                kode = normalize_str(kode_raw)
                if not kode:
                    raise ValueError("kode dosen kosong")

                nama = str(nama_raw).strip() if nama_raw else None
                if not nama:
                    raise ValueError("nama dosen kosong")

                # nidn / nip — tolerant: strip whitespace, None jika kosong
                nidn = _clean_nullable_str(nidn_raw, max_len=20)
                nip = _clean_nullable_str(nip_raw, max_len=25)

                jabfung = str(jabfung_raw).strip() if jabfung_raw else None
                kjfd = str(kjfd_raw).strip() if kjfd_raw else None

                # Resolve homebase prodi — None jika tidak ditemukan (tidak raise)
                homebase_prodi_id = None
                if homebase_raw:
                    homebase_norm = normalize_str(homebase_raw)
                    prodi = self.db_session.query(Prodi).filter(
                        Prodi.nama.ilike(f"%{homebase_norm}%")
                    ).first()
                    if prodi:
                        homebase_prodi_id = prodi.id
                    else:
                        # Coba match via strata+nama (misal "S-1 Matematika")
                        parts = homebase_norm.split()
                        if len(parts) >= 2:
                            nama_part = " ".join(parts[1:])
                            prodi = self.db_session.query(Prodi).filter(
                                Prodi.nama.ilike(f"%{nama_part}%")
                            ).first()
                            if prodi:
                                homebase_prodi_id = prodi.id

                # Parse tgl_lahir — tolerant
                tgl_lahir: Optional[date] = None
                if tgl_lahir_raw is not None:
                    if isinstance(tgl_lahir_raw, datetime):
                        tgl_lahir = tgl_lahir_raw.date()
                    elif isinstance(tgl_lahir_raw, date):
                        tgl_lahir = tgl_lahir_raw
                    else:
                        try:
                            tgl_lahir = datetime.strptime(
                                str(tgl_lahir_raw).strip(), "%Y-%m-%d"
                            ).date()
                        except ValueError:
                            pass  # tgl_lahir tetap None

                # Upsert berdasarkan kode
                existing = self.db_session.query(Dosen).filter(
                    Dosen.kode == kode.upper()
                ).first()

                if existing:
                    existing.nama = nama
                    existing.jabfung = jabfung
                    existing.kjfd = kjfd
                    existing.homebase_prodi_id = homebase_prodi_id
                    existing.tgl_lahir = tgl_lahir
                    # nidn/nip: update hanya jika nilai baru tidak None
                    # (hindari overwrite data valid dengan None)
                    if nidn is not None:
                        existing.nidn = nidn
                    if nip is not None:
                        existing.nip = nip
                    updated += 1
                else:
                    self.db_session.add(Dosen(
                        kode=kode.upper(),
                        nama=nama,
                        nidn=nidn,
                        nip=nip,
                        jabfung=jabfung,
                        kjfd=kjfd,
                        homebase_prodi_id=homebase_prodi_id,
                        tgl_lahir=tgl_lahir,
                    ))
                    inserted += 1

                self.db_session.flush()

            except Exception as exc:
                skipped += 1
                warnings.append(ImportWarning(
                    row=row_num, sheet=SHEET, value=row, reason=str(exc)
                ))
                logger.warning("_import_dosen row %d: %s", row_num, exc)

        self.db_session.commit()
        return ImportResult(total=total, inserted=inserted, updated=updated,
                            skipped=skipped, warnings=warnings)

    def import_mata_kuliah_kelas(self, file) -> ImportResult:
        """
        Import data kelas paralel mata kuliah dari file db_mata_kuliah.xlsx.

        Membaca sheet 'db_kelas' dan membuat/memperbarui record MataKuliahKelas.
        Baris di mana kode_mk tidak ditemukan di tabel mata_kuliah dilewati
        dengan warning (bukan error) — tolerant import strategy.

        Kolom sheet db_kelas:
          [0] kode_mk    → lookup MataKuliah (case-insensitive)
          [1] mata_kuliah → (ignored, sudah ada di MataKuliah)
          [2] kelas      → nullable (A / B / C / None)
          [3] mk_kelas   → label (NOT NULL)
          [4..9]         → (ignored)
          [10] ket       → nullable keterangan

        Upsert berdasarkan (mata_kuliah_id, kelas).

        Args:
            file: File-like object atau path ke db_mata_kuliah.xlsx.

        Returns:
            ImportResult dengan ringkasan total, inserted, updated, skipped,
            dan daftar warnings per baris bermasalah.
        """
        SHEET = "db_kelas"

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception as exc:
            logger.error("import_mata_kuliah_kelas: gagal membuka file Excel: %s", exc)
            return ImportResult(
                total=0, inserted=0, updated=0, skipped=1,
                warnings=[ImportWarning(row=0, sheet="(file)", value=str(file), reason=str(exc))],
            )

        if SHEET not in wb.sheetnames:
            logger.warning("import_mata_kuliah_kelas: sheet '%s' tidak ditemukan", SHEET)
            return ImportResult(total=0, inserted=0, updated=0, skipped=0)

        ws = wb[SHEET]
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
        total = inserted = updated = skipped = 0
        warnings: list[ImportWarning] = []

        for row_num, row in enumerate(rows, start=2):
            total += 1
            try:
                kode_mk_raw = row[0] if len(row) > 0 else None
                kelas_raw = row[2] if len(row) > 2 else None
                label_raw = row[3] if len(row) > 3 else None
                ket_raw = row[10] if len(row) > 10 else None

                kode_mk = normalize_str(kode_mk_raw)
                if not kode_mk:
                    raise ValueError("kode_mk kosong")

                label = str(label_raw).strip() if label_raw else None
                if not label:
                    raise ValueError("label (mk_kelas) kosong")

                # kelas: nullable — None jika kosong
                kelas: Optional[str] = None
                if kelas_raw is not None:
                    kelas_str = str(kelas_raw).strip()
                    kelas = kelas_str[:5] if kelas_str else None

                ket = _clean_nullable_str(ket_raw)

                # Lookup MataKuliah — FK miss = warning, bukan error
                mk = self.db_session.query(MataKuliah).filter(
                    MataKuliah.kode.ilike(kode_mk)
                ).first()
                if mk is None:
                    skipped += 1
                    warnings.append(ImportWarning(
                        row=row_num, sheet=SHEET, value=kode_mk_raw,
                        reason=f"kode_mk '{kode_mk_raw}' tidak ditemukan di tabel mata_kuliah",
                    ))
                    logger.warning(
                        "import_mata_kuliah_kelas row %d: kode_mk '%s' tidak ditemukan",
                        row_num, kode_mk_raw,
                    )
                    continue

                # Upsert berdasarkan (mata_kuliah_id, kelas)
                existing = self.db_session.query(MataKuliahKelas).filter(
                    MataKuliahKelas.mata_kuliah_id == mk.id,
                    MataKuliahKelas.kelas == kelas,
                ).first()

                if existing:
                    existing.label = label
                    existing.ket = ket
                    updated += 1
                else:
                    self.db_session.add(MataKuliahKelas(
                        mata_kuliah_id=mk.id,
                        kelas=kelas,
                        label=label,
                        ket=ket,
                    ))
                    inserted += 1

                self.db_session.flush()

            except Exception as exc:
                skipped += 1
                warnings.append(ImportWarning(
                    row=row_num, sheet=SHEET, value=row, reason=str(exc)
                ))
                logger.warning("import_mata_kuliah_kelas row %d: %s", row_num, exc)

        self.db_session.commit()
        return ImportResult(total=total, inserted=inserted, updated=updated,
                            skipped=skipped, warnings=warnings)

    # ------------------------------------------------------------------
    # Timeslot sesi mapping helpers
    # ------------------------------------------------------------------

    # Map jam_mulai prefix → sesi number (sesuai seed data)
    # Sesi 1: 07:30, Sesi 2: 10:00, Sesi 3: 13:00
    _WAKTU_TO_SESI: dict[str, int] = {
        "07:": 1,
        "08:": 1,
        "09:": 1,
        "10:": 2,
        "11:": 2,
        "12:": 2,
        "13:": 3,
        "14:": 3,
        "15:": 3,
    }

    # Map hari name (lowercase) → kode prefix used in timeslot.kode
    _HARI_TO_KODE: dict[str, str] = {
        "senin": "mon",
        "selasa": "tue",
        "rabu": "wed",
        "kamis": "thu",
        "jumat": "fri",
    }

    def _resolve_timeslot(self, hari_raw, waktu_raw) -> Optional["Timeslot"]:
        """
        Resolve Timeslot dari nilai Hari dan Waktu di Excel.

        Strategi:
        1. Normalise hari → kode prefix (mon/tue/wed/thu/fri)
        2. Derive sesi dari jam mulai (07:xx → s1, 10:xx → s2, 13:xx → s3)
        3. Lookup Timeslot by kode = "{prefix}_s{sesi}"

        Returns None jika tidak dapat di-resolve.
        """
        hari_norm = normalize_str(hari_raw)
        kode_prefix = self._HARI_TO_KODE.get(hari_norm)
        if not kode_prefix:
            return None

        waktu_str = str(waktu_raw).strip() if waktu_raw else ""
        # Ambil jam mulai (sebelum '-')
        jam_mulai = waktu_str.split("-")[0].strip() if "-" in waktu_str else waktu_str
        # Ambil 3 karakter pertama untuk prefix matching (e.g. "07:", "10:", "13:")
        jam_prefix = jam_mulai[:3] if len(jam_mulai) >= 3 else ""
        sesi = self._WAKTU_TO_SESI.get(jam_prefix)
        if not sesi:
            return None

        kode = f"{kode_prefix}_s{sesi}"
        return self.db_session.query(Timeslot).filter(
            Timeslot.kode == kode
        ).first()

    def _resolve_mk_kelas(self, kode_raw, kelas_raw) -> Optional["MataKuliahKelas"]:
        """
        Resolve MataKuliahKelas dari kode MK dan huruf kelas.

        Mencoba kedua kolom kode (kurikulum 2021 dan 2025).
        kelas_raw boleh None (untuk MK tanpa kelas paralel).

        Returns None jika tidak ditemukan.
        """
        kode = normalize_str(kode_raw)
        if not kode:
            return None

        kelas: Optional[str] = None
        if kelas_raw is not None:
            kelas_str = str(kelas_raw).strip()
            kelas = kelas_str if kelas_str else None

        # Lookup MataKuliah by kode (case-insensitive)
        mk = self.db_session.query(MataKuliah).filter(
            MataKuliah.kode.ilike(kode)
        ).first()
        if mk is None:
            return None

        # Lookup MataKuliahKelas by (mata_kuliah_id, kelas)
        mk_kelas = self.db_session.query(MataKuliahKelas).filter(
            MataKuliahKelas.mata_kuliah_id == mk.id,
            MataKuliahKelas.kelas == kelas,
        ).first()
        return mk_kelas

    def _resolve_ruang(self, ruang_raw) -> Optional["Ruang"]:
        """
        Resolve Ruang dari nilai kolom Ruang di Excel.

        Returns None jika kolom kosong atau ruang tidak ditemukan di DB.
        ruang_id adalah NULLABLE — tidak wajib ada.
        """
        if ruang_raw is None:
            return None
        nama = str(ruang_raw).strip()
        if not nama:
            return None
        return self.db_session.query(Ruang).filter(
            Ruang.nama.ilike(nama)
        ).first()

    def import_jadwal(self, file, sesi_id: str) -> ImportResult:
        """
        Import assignment jadwal dari file Excel historis.

        Membaca sheet jadwal dari file Excel semester sebelumnya dan membuat
        JadwalAssignment yang terhubung ke SesiJadwal yang ditentukan.

        Struktur sheet yang didukung (header di baris 9):
          col[1]  Hari
          col[3]  Ruang
          col[4]  Waktu  (e.g. "07:30-10:00")
          col[7]  Kode MK kurikulum 2021
          col[8]  Kode MK kurikulum 2025
          col[12] Kelas  (A/B/C/None)
          col[17] Dosen Pengajar I
          col[18] Dosen Pengajar II (opsional)

        Upsert: skip jika (sesi_id, mk_kelas_id) sudah ada.

        Args:
            file: File-like object atau path ke file Excel jadwal historis.
            sesi_id: UUID string dari SesiJadwal target di database.

        Returns:
            ImportResult dengan ringkasan hasil import jadwal.
        """
        import uuid as _uuid

        VALID_HARI = {"senin", "selasa", "rabu", "kamis", "jumat"}
        # Sheet name patterns to try (most recent format first)
        SHEET_CANDIDATES = [
            "Jadwal Genap 2025 2026",
            "Jadwal Ganjil 2025 2026",
            "Jadwal Genap 2024 2025",
            "Jadwal Ganjil 2024 2025",
            "Jadwal Genap 2023 2024",
        ]

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        except Exception as exc:
            logger.error("import_jadwal: gagal membuka file Excel: %s", exc)
            return ImportResult(
                total=0, inserted=0, updated=0, skipped=1,
                warnings=[ImportWarning(row=0, sheet="(file)", value=str(file), reason=str(exc))],
            )

        # Find the schedule sheet
        ws = None
        sheet_name = None
        for candidate in SHEET_CANDIDATES:
            if candidate in wb.sheetnames:
                ws = wb[candidate]
                sheet_name = candidate
                break
        # Fallback: use first sheet that contains "Jadwal" in name
        if ws is None:
            for name in wb.sheetnames:
                if "jadwal" in name.lower():
                    ws = wb[name]
                    sheet_name = name
                    break
        if ws is None:
            logger.error("import_jadwal: tidak ada sheet jadwal ditemukan. Sheets: %s", wb.sheetnames)
            return ImportResult(
                total=0, inserted=0, updated=0, skipped=1,
                warnings=[ImportWarning(
                    row=0, sheet="(file)", value=str(wb.sheetnames),
                    reason="Sheet jadwal tidak ditemukan di file Excel",
                )],
            )

        # Parse sesi_id
        try:
            sesi_uuid = _uuid.UUID(str(sesi_id))
        except (ValueError, AttributeError) as exc:
            return ImportResult(
                total=0, inserted=0, updated=0, skipped=1,
                warnings=[ImportWarning(row=0, sheet=sheet_name, value=sesi_id, reason=f"sesi_id tidak valid: {exc}")],
            )

        total = inserted = updated = skipped = 0
        warnings: list[ImportWarning] = []

        # Find header row — look for row containing 'Hari' in col[1]
        header_row = 9  # default based on known file structure
        all_rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
        for i, row in enumerate(all_rows, start=1):
            if row and len(row) > 1 and normalize_str(row[1]) in ("hari", "hari  "):
                header_row = i
                break

        data_rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))

        for row_num, row in enumerate(data_rows, start=header_row + 1):
            # Skip rows with no hari or non-valid hari
            if not row or len(row) < 18:
                continue
            hari_raw = row[1]
            if normalize_str(hari_raw) not in VALID_HARI:
                continue

            total += 1
            try:
                waktu_raw = row[4]
                kode_2021 = row[7]
                kode_2025 = row[8]
                kelas_raw = row[12]
                dosen1_raw = row[17]
                dosen2_raw = row[18] if len(row) > 18 else None
                ruang_raw = row[3]

                # --- Resolve timeslot ---
                timeslot = self._resolve_timeslot(hari_raw, waktu_raw)
                if timeslot is None:
                    skipped += 1
                    warnings.append(ImportWarning(
                        row=row_num, sheet=sheet_name,
                        value=f"hari={hari_raw!r} waktu={waktu_raw!r}",
                        reason="Timeslot tidak dapat di-resolve dari hari/waktu",
                    ))
                    continue

                # --- Resolve mk_kelas (try both kode columns) ---
                mk_kelas = None
                for kode_raw in (kode_2025, kode_2021):
                    if kode_raw:
                        mk_kelas = self._resolve_mk_kelas(kode_raw, kelas_raw)
                        if mk_kelas:
                            break

                if mk_kelas is None:
                    skipped += 1
                    warnings.append(ImportWarning(
                        row=row_num, sheet=sheet_name,
                        value=f"kode_2021={kode_2021!r} kode_2025={kode_2025!r} kelas={kelas_raw!r}",
                        reason="MataKuliahKelas tidak ditemukan untuk kode/kelas ini",
                    ))
                    continue

                # --- Resolve dosen1 (required) ---
                dosen1 = resolve_dosen(str(dosen1_raw).strip() if dosen1_raw else "", self.db_session)
                if dosen1 is None:
                    skipped += 1
                    warnings.append(ImportWarning(
                        row=row_num, sheet=sheet_name,
                        value=dosen1_raw,
                        reason=f"Dosen I '{dosen1_raw}' tidak ditemukan di database",
                    ))
                    continue

                # --- Resolve dosen2 (optional) ---
                dosen2 = None
                if dosen2_raw:
                    dosen2_str = str(dosen2_raw).strip()
                    if dosen2_str:
                        dosen2 = resolve_dosen(dosen2_str, self.db_session)
                        if dosen2 is None:
                            warnings.append(ImportWarning(
                                row=row_num, sheet=sheet_name,
                                value=dosen2_raw,
                                reason=f"Dosen II '{dosen2_raw}' tidak ditemukan — dosen2_id diset None",
                            ))

                # --- Resolve ruang (optional, nullable) ---
                ruang = self._resolve_ruang(ruang_raw)
                ruang_id = ruang.id if ruang else None

                # --- Upsert: skip if (sesi_id, mk_kelas_id) already exists ---
                existing = self.db_session.query(JadwalAssignment).filter(
                    JadwalAssignment.sesi_id == sesi_uuid,
                    JadwalAssignment.mk_kelas_id == mk_kelas.id,
                ).first()

                if existing:
                    updated += 1
                else:
                    self.db_session.add(JadwalAssignment(
                        sesi_id=sesi_uuid,
                        mk_kelas_id=mk_kelas.id,
                        dosen1_id=dosen1.id,
                        dosen2_id=dosen2.id if dosen2 else None,
                        timeslot_id=timeslot.id,
                        ruang_id=ruang_id,
                    ))
                    inserted += 1

                self.db_session.flush()

            except Exception as exc:
                skipped += 1
                warnings.append(ImportWarning(
                    row=row_num, sheet=sheet_name, value=row, reason=str(exc)
                ))
                logger.warning("import_jadwal row %d: %s", row_num, exc)

        self.db_session.commit()
        return ImportResult(total=total, inserted=inserted, updated=updated,
                            skipped=skipped, warnings=warnings)
